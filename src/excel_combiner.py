import re
import os
import tempfile
from typing import Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from .config import Config


class ExcelCombiner:
    def __init__(self, output_filename: str = None):
        self.output_filename = output_filename or Config.OUTPUT_FILENAME
        self.workbook = Workbook()
        self.used_sheet_names = set()  # Track used names for uniqueness
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize and truncate sheet name to meet Excel requirements"""
        # Remove invalid characters: \ / ? * [ ] :
        invalid_chars = r'[\\/*?\[\]:]'
        sanitized = re.sub(invalid_chars, '_', name)
        
        # Cannot be "History" (case insensitive)
        if sanitized.lower() == 'history':
            sanitized = 'History_Sheet'
        
        # Truncate to 31 characters
        if len(sanitized) > 31:
            sanitized = sanitized[:31]
        
        # Ensure uniqueness
        original_sanitized = sanitized
        counter = 1
        while sanitized in self.used_sheet_names:
            # If duplicate, add counter (accounting for character limit)
            suffix = f"_{counter}"
            max_base_length = 31 - len(suffix)
            sanitized = original_sanitized[:max_base_length] + suffix
            counter += 1
        
        self.used_sheet_names.add(sanitized)
        return sanitized
    
    def add_dataframe(self, df: pd.DataFrame, sheet_name: str):
        """Add a DataFrame as a new worksheet"""
        # Sanitize the sheet name for Excel compatibility
        sanitized_name = self._sanitize_sheet_name(sheet_name)
        
        # Log name changes for user awareness
        if sanitized_name != sheet_name:
            print(f"📝 Sheet name changed: '{sheet_name}' → '{sanitized_name}'")
            print(f"   (Excel limit: 31 chars, invalid chars removed)")
        
        # Create new worksheet
        worksheet = self.workbook.create_sheet(title=sanitized_name)
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(r)
        
        # Format headers
        if worksheet.max_row > 0:
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCCCCC",
                    end_color="CCCCCC",
                    fill_type="solid"
                )
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def save(self, target_path: str = None):
        """Save the workbook to file with UNC path support"""
        if not self.workbook.worksheets:
            print("No data to save!")
            return False
        
        # Use provided path or the configured path
        save_path = target_path or self.output_filename
        
        try:
            # For UNC paths, we might need special handling
            if save_path.startswith("\\\\"):
                # This is a UNC path, let the UNC manager handle it
                return self._save_with_unc_support(save_path)
            else:
                # Standard local save
                return self._save_locally(save_path)
                
        except Exception as e:
            print(f"Error saving file: {e}")
            return False
    
    def _save_locally(self, file_path: str) -> bool:
        """Save file locally with directory creation"""
        try:
            # Ensure directory exists
            directory = os.path.dirname(file_path)
            if directory and not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
            
            self.workbook.save(file_path)
            print(f"✅ Successfully saved: {file_path}")
            print(f"📊 Total sheets: {len(self.workbook.worksheets)}")
            return True
            
        except Exception as e:
            print(f"❌ Error saving locally: {e}")
            return False
    
    def _save_with_unc_support(self, unc_path: str) -> bool:
        """Save file to UNC path using UNC path manager"""
        try:
            # Import here to avoid circular imports
            from .unc_path_manager import unc_path_manager
            
            # Create a temporary file first
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_path = temp_file.name
            
            # Save to temporary file
            self.workbook.save(temp_path)
            print(f"📝 Created temporary file: {temp_path}")
            
            # Use UNC manager to move to final location
            success = unc_path_manager.save_file_safely(temp_path, unc_path)
            
            # Clean up temporary file
            try:
                os.unlink(temp_path)
            except Exception as e:
                print(f"⚠️  Could not remove temporary file: {e}")
            
            if success:
                print(f"📊 Total sheets: {len(self.workbook.worksheets)}")
                return True
            else:
                return False
                
        except Exception as e:
            print(f"❌ Error saving with UNC support: {e}")
            return False
    
    def get_sheet_summary(self) -> Dict[str, int]:
        """Get summary of sheets and their row counts"""
        summary = {}
        for sheet in self.workbook.worksheets:
            summary[sheet.title] = sheet.max_row - 1  # Subtract header row
        return summary
